
# ─────────────────────────────────────────────────────────────────────────────
# ⚠️  DEPRECATED — 此脚本已被废弃
#
# 废弃时间：2026-06-09
# 替代版本：见 scripts/DEPRECATED.md
#
# 如需使用，请改用对应的新版本脚本。
# 本文件仅保留用于历史参考，不建议在实际研究流程中使用。
# ─────────────────────────────────────────────────────────────────────────────
# [原文件内容继续]

#!/usr/bin/env python3
"""
批量获取指定人员负责股票的MSCI ESG评级。

用法：
    python scripts/fetch_msci_cyh.py --person "陈宇浩" --excel /path/to/input.xlsx

环境变量：
    EXCEL_PATH   — 输入Excel文件路径
    PERSON_NAME  — Excel中负责人的列值
"""
import argparse
import json
import os
import ssl
import time
import urllib.request
from pathlib import Path

import openpyxl

ctx = ssl.create_default_context()
ctx.check_hostname = False
ctx.verify_mode = ssl.CERT_NONE

HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36',
    'Referer': 'https://finance.sina.com.cn/esg/grade.shtml',
}

DEFAULT_EXCEL_PATH = str(Path.home() / 'Desktop' / '2026MSCI级别6.1(1).xlsx')
DEFAULT_OUTPUT_PATH = str(Path.home() / 'Desktop' / '2026MSCI级别6.1(1)_已填充.xlsx')

def get_msci(symbol):
    url = f'https://global.finance.sina.com.cn/api/openapi.php/EsgService.getEsgStockInfo?symbol={symbol}'
    try:
        req = urllib.request.Request(url, headers=HEADERS)
        with urllib.request.urlopen(req, timeout=15, context=ctx) as response:
            data = json.loads(response.read().decode('utf-8'))
            for info in data.get('result', {}).get('data', {}).get('info', []):
                if info.get('agency_name') == 'MSCI':
                    return info.get('esg_score'), info.get('esg_dt')
    except Exception:
        pass
    return None, None

def main():
    parser = argparse.ArgumentParser(description="批量获取指定人员的MSCI ESG评级")
    parser.add_argument('--person', default=os.environ.get('PERSON_NAME', '陈宇浩'),
                        help='Excel中负责人的列值（环境变量: PERSON_NAME）')
    parser.add_argument('--excel', default=os.environ.get('EXCEL_PATH', DEFAULT_EXCEL_PATH),
                        help='输入Excel路径（环境变量: EXCEL_PATH）')
    parser.add_argument('--output', default=DEFAULT_OUTPUT_PATH,
                        help='输出Excel路径')
    args = parser.parse_args()

    person = args.person
    excel = args.excel
    output = args.output
    cache_key = person.replace(' ', '_')
    cache_file = str(Path.home() / 'Desktop' / f'论文-研报工作流/data/msci_esg_ratings_{cache_key}.json')

    print("1. 读取Excel...")
    wb = openpyxl.load_workbook(excel)
    ws = wb['2025年度上市公司名单']

    target_stocks = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if row[6] == person:
            ts_code = row[1]
            if ts_code and ts_code.endswith('.SH'):
                symbol = 'sh' + ts_code.replace('.SH', '')
            elif ts_code and ts_code.endswith('.SZ'):
                symbol = 'sz' + ts_code.replace('.SZ', '')
            else:
                symbol = None
            target_stocks.append({
                'row_idx': row_idx,
                'name': row[0],
                'ts_code': ts_code,
                'hk_code': row[2],
                'short': row[3],
                'symbol': symbol,
                'rating': row[4],
            })

    unfilled = [s for s in target_stocks if s['rating'] is None]
    print(f"   [{person}] 总数: {len(target_stocks)}")
    print(f"   已填充: {sum(1 for s in target_stocks if s['rating'] is not None)}")
    print(f"   未填充: {len(unfilled)}")

    cache = {}
    if os.path.exists(cache_file):
        with open(cache_file, encoding='utf-8') as f:
            cache = json.load(f)
        print(f"   缓存已有: {len(cache)} 条")

    # 复用其他人缓存
    for other_cache in Path.home().glob('Desktop/论文-研报工作流/data/msci_esg_ratings_*.json'):
        if other_cache.name != cache_file:
            try:
                with open(other_cache, encoding='utf-8') as f:
                    other_data = json.load(f)
                reused = sum(1 for s in unfilled if s['symbol'] and s['symbol'].upper() in other_data)
                if reused > 0:
                    print(f"   复用 [{other_cache.name}]: {reused} 条")
                    for s in unfilled:
                        if s['symbol'] and s['symbol'].upper() in other_data:
                            cache[s['symbol'].upper()] = other_data[s['symbol'].upper()]
            except Exception:
                pass

    to_fetch = []
    for s in unfilled:
        if s['symbol']:
            sym_up = s['symbol'].upper()
            if sym_up not in cache:
                to_fetch.append(s)

    print(f"   需新获取: {len(to_fetch)} 只")

    if to_fetch:
        print("\n2. 开始获取...")
        fetched = 0
        start = time.time()
        for i, s in enumerate(to_fetch):
            sym = s['symbol']
            rating, date = get_msci(sym)
            cache[sym.upper()] = {'msci_rating': rating, 'msci_date': date}
            if rating:
                fetched += 1
            if (i+1) % 25 == 0:
                elapsed = time.time() - start
                rate = elapsed / (i+1)
                eta = rate * (len(to_fetch) - i - 1) if (i+1) > 0 else 0
                print(f"   {i+1}/{len(to_fetch)} | 成功: {fetched} | ETA: {eta/60:.1f}min")
            time.sleep(0.3)

        with open(cache_file, 'w', encoding='utf-8') as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
        print(f"   缓存已保存, 新增成功: {fetched}")

    print("\n3. 回填Excel...")
    filled = 0
    no_rating = 0

    for s in target_stocks:
        row_idx = s['row_idx']
        if s['rating'] is not None:
            continue
        if not s['symbol']:
            no_rating += 1
            continue
        sym_upper = s['symbol'].upper()
        if sym_upper in cache:
            result = cache[sym_upper]
            rating = result.get('msci_rating')
            date = result.get('msci_date')
            if rating and rating != '-':
                ws.cell(row=row_idx, column=5, value=rating)
                ws.cell(row=row_idx, column=6, value=date)
                filled += 1
            else:
                no_rating += 1
        else:
            no_rating += 1

    wb.save(output)
    print(f"   保存: {output}")
    total_filled = sum(1 for s in target_stocks if s['rating'] is not None or (s['symbol'] and cache.get(s['symbol'].upper(),{}).get('msci_rating')))
    print(f"   本次新增填充: {filled}")
    print(f"   [{person}] 总计已填充: {total_filled}/{len(target_stocks)}")

    dist = {}
    for s in target_stocks:
        if s['symbol']:
            sym_up = s['symbol'].upper()
            r = s['rating'] if s['rating'] else cache.get(sym_up, {}).get('msci_rating')
            if r and r != '-':
                dist[r] = dist.get(r, 0) + 1

    print("\n=== 评级分布 ===")
    for r in sorted(dist.keys()):
        print(f"  {r}: {dist[r]}")
    print("Done!")

if __name__ == '__main__':
    main()
